"""
=============================================================================
🎰 تطبيق اليانصيب الأردني المتطور - النسخة 3.0
=============================================================================
تطبيق شامل لتوليد وتحليل أرقام اليانصيب مع تحليلات متقدمة

المطور: محمد العمري
التاريخ: فبراير 2026
الإصدار: 3.0.0

التحسينات الجديدة:
- الوضع الداكن افتراضياً
- تحميل البيانات تلقائياً من الملف المحلي
- نافذة ترشيح متقدمة للتذاكر (6-10 أرقام)
=============================================================================
"""

import streamlit as st
import pandas as pd
import numpy as np
import random
import time
import logging
import io
import os
import requests  # للتحميل من GitHub
from datetime import datetime, timedelta
from collections import Counter, defaultdict
from typing import List, Dict, Optional, Tuple, Set, Union
from itertools import chain, combinations

# رسوم بيانية
import plotly.express as px
import plotly.graph_objects as go

# تصدير Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# تصدير PDF
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    PDF_AVAILABLE = True
except:
    PDF_AVAILABLE = False

import warnings
warnings.filterwarnings('ignore')

# ==============================================================================
# 1. إعدادات النظام
# ==============================================================================

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s', datefmt='%H:%M:%S')
logger = logging.getLogger("JordanLottery")

class LotteryConfig:
    MIN_NUM = 1
    MAX_NUM = 32
    DEFAULT_TICKET_SIZE = 6
    MIN_TICKET_SIZE = 6
    MAX_TICKET_SIZE = 10
    MAX_GENERATION_ATTEMPTS = 50000
    STRICT_SHADOW_ATTEMPTS = 15000
    DEFAULT_SUM_TOLERANCE = 0.15
    MAX_BATCH_SIZE = 10
    
    # ملف البيانات في نفس المجلد
    DATA_FILE = "history.xlsx"
    
    TICKET_PRICES = {6: 1, 7: 7, 8: 28, 9: 84, 10: 210}
    MATCH_PRIZES = {3: 1, 4: 15, 5: 500, 6: "JACKPOT"}

def initialize_session_state():
    defaults = {
        'history_df': None,
        'analyzer': None,
        'generator': None,
        'last_result': None,
        'theme': 'dark',  # الوضع الداكن افتراضياً
        'hot_color': '#22c55e',
        'cold_color': '#3b82f6',
        'balanced_color': '#f59e0b',
        'generated_tickets_session': [],
        'data_loaded_automatically': False,
        'filtered_tickets': [],
        'auto_loaded': False,  # للتحميل التلقائي من GitHub
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

# ==============================================================================
# 2. الثيم والتنسيقات
# ==============================================================================

def apply_theme():
    theme = st.session_state.theme
    
    if theme == 'dark':
        bg_color = '#1e1e1e'
        text_color = '#ffffff'
        card_bg = '#2d2d2d'
        border_color = '#404040'
    else:
        bg_color = '#ffffff'
        text_color = '#1f2937'
        card_bg = '#f9fafb'
        border_color = '#e5e7eb'
    
    st.markdown(f"""
    <style>
        .stApp {{
            background-color: {bg_color};
            color: {text_color};
        }}
        
        @keyframes fadeIn {{
            from {{ opacity: 0; transform: translateY(10px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}
        
        .number-animated {{
            animation: fadeIn 0.4s ease-out;
            display: inline-block;
            margin: 3px;
        }}
        
        .lottery-number {{
            display: inline-block;
            background: {st.session_state.hot_color};
            color: white;
            padding: 8px 14px;
            margin: 3px;
            border-radius: 50%;
            font-weight: bold;
            border: 2px solid rgba(255,255,255,0.3);
            box-shadow: 0 2px 5px rgba(0,0,0,0.2);
        }}
        
        .filter-card {{
            background: {card_bg};
            border: 2px solid {border_color};
            border-radius: 10px;
            padding: 20px;
            margin: 10px 0;
        }}
        
        .score-badge {{
            display: inline-block;
            padding: 5px 12px;
            border-radius: 15px;
            font-weight: bold;
            margin: 0 5px;
        }}
        
        .footer {{
            position: fixed;
            bottom: 0;
            left: 0;
            width: 100%;
            background-color: {card_bg};
            color: {text_color};
            text-align: center;
            padding: 10px;
            font-size: 14px;
            border-top: 1px solid {border_color};
            z-index: 999;
        }}
    </style>
    """, unsafe_allow_html=True)

# ==============================================================================
# 3. تحميل البيانات
# ==============================================================================

@st.cache_data(show_spinner=False)
def load_data_automatically():
    """تحميل البيانات تلقائياً من الملف المحلي"""
    try:
        # البحث عن الملف في المجلد الحالي
        if os.path.exists(LotteryConfig.DATA_FILE):
            df, msg = load_and_process_data(LotteryConfig.DATA_FILE)
            return df, msg
        else:
            return None, f"الملف {LotteryConfig.DATA_FILE} غير موجود"
    except Exception as e:
        logger.error(f"Error loading data: {e}")
        return None, f"خطأ في تحميل البيانات: {str(e)}"

@st.cache_data(show_spinner=False, ttl=300)  # تحديث كل 5 دقائق
def load_from_github(url: str = None) -> Tuple[Optional[pd.DataFrame], str]:
    """تحميل البيانات من GitHub مع تحديث تلقائي"""
    import requests
    from io import BytesIO
    
    try:
        if url is None:
            # استخدام اسم الملف الأحدث
            url = "https://raw.githubusercontent.com/MohamedOmariJo/omari/main/250.xlsx"
        
        # تحميل الملف من GitHub
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        # قراءة الملف من الذاكرة
        file_content = BytesIO(response.content)
        df = pd.read_excel(file_content)
        
        # تنظيف أولي
        df.dropna(how='all', inplace=True)
        
        # التحقق من الأعمدة
        required_cols = ['N1', 'N2', 'N3', 'N4', 'N5', 'N6']
        if not set(required_cols).issubset(df.columns):
            return None, "خطأ: الملف لا يحتوي على أعمدة الأرقام (N1...N6)"

        # تحويل الأرقام
        for col in required_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        df.dropna(subset=required_cols, inplace=True)
        df['numbers'] = df[required_cols].values.tolist()
        
        # فلتر النطاق (1-32)
        def is_valid_draw(nums):
            return all(LotteryConfig.MIN_NUM <= int(n) <= LotteryConfig.MAX_NUM for n in nums)

        df = df[df['numbers'].apply(is_valid_draw)]
        
        if df.empty:
            return None, "خطأ: لا توجد بيانات صالحة"

        # ترتيب الأرقام
        df['numbers'] = df['numbers'].apply(lambda x: sorted([int(n) for n in x]))
        
        # توحيد عمود المعرف
        if 'رقم السحب' in df.columns:
            df = df.rename(columns={'رقم السحب': 'draw_id'})
        elif 'DrawID' in df.columns:
            df = df.rename(columns={'DrawID': 'draw_id'})
        elif 'draw_id' not in df.columns:
            df['draw_id'] = range(1, len(df) + 1)
        
        # التاريخ
        if 'تاريخ السحب' in df.columns:
            df['date'] = pd.to_datetime(df['تاريخ السحب'], errors='coerce')
        elif 'date' not in df.columns:
            start_date = datetime(2023, 9, 17)
            df['date'] = [start_date + timedelta(days=i*3) for i in range(len(df))]
        
        # إضافة الحقول المحسوبة
        df['sum'] = df['numbers'].apply(sum)
        df['odd_count'] = df['numbers'].apply(lambda x: sum(1 for n in x if n % 2 == 1))
        df['even_count'] = df['numbers'].apply(lambda x: sum(1 for n in x if n % 2 == 0))
        
        return df, f"✅ تم تحميل {len(df)} سحب من GitHub"
        
    except requests.exceptions.RequestException as e:
        logger.error(f"GitHub loading error: {e}")
        return None, f"خطأ في الاتصال بـ GitHub: {str(e)}"
    except Exception as e:
        logger.error(f"Data processing error: {e}")
        return None, f"خطأ في معالجة الملف: {str(e)}"


@st.cache_data(show_spinner=False)
def load_and_process_data(file_input):
    try:
        is_csv = False
        if isinstance(file_input, str):
            is_csv = file_input.endswith('.csv')
        elif hasattr(file_input, 'name'):
            is_csv = file_input.name.endswith('.csv')

        if is_csv:
            df = pd.read_csv(file_input)
        else:
            df = pd.read_excel(file_input)
        
        df.dropna(how='all', inplace=True)
        
        required_cols = ['N1', 'N2', 'N3', 'N4', 'N5', 'N6']
        if not set(required_cols).issubset(df.columns):
            return None, "خطأ: الملف لا يحتوي على أعمدة الأرقام (N1...N6)"

        for col in required_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        df.dropna(subset=required_cols, inplace=True)
        df['numbers'] = df[required_cols].values.tolist()
        
        def is_valid_draw(nums):
            return all(LotteryConfig.MIN_NUM <= int(n) <= LotteryConfig.MAX_NUM for n in nums)

        df = df[df['numbers'].apply(is_valid_draw)]
        
        if df.empty:
            return None, "خطأ: لا توجد بيانات صالحة"

        df['numbers'] = df['numbers'].apply(lambda x: sorted([int(n) for n in x]))
        
        if 'رقم السحب' in df.columns:
            df = df.rename(columns={'رقم السحب': 'draw_id'})
        elif 'DrawID' in df.columns:
            df = df.rename(columns={'DrawID': 'draw_id'})
        elif 'draw_id' not in df.columns:
            df['draw_id'] = range(1, len(df) + 1)
        
        if 'تاريخ السحب' in df.columns:
            df['date'] = pd.to_datetime(df['تاريخ السحب'], errors='coerce')
        elif 'date' not in df.columns:
            start_date = datetime(2023, 9, 17)
            df['date'] = [start_date + timedelta(days=i*3) for i in range(len(df))]
        
        df['sum'] = df['numbers'].apply(sum)
        df['odd_count'] = df['numbers'].apply(lambda x: sum(1 for n in x if n % 2 == 1))
        df['even_count'] = df['numbers'].apply(lambda x: sum(1 for n in x if n % 2 == 0))
        
        return df, "Success"
        
    except Exception as e:
        logger.error(f"Error: {e}")
        return None, f"خطأ: {str(e)}"

# ==============================================================================
# 4. المحلل المتقدم
# ==============================================================================

class AdvancedAnalyzer:
    def __init__(self, history_df: pd.DataFrame):
        self.history_df = history_df
        self.past_draws_sets = [set(nums) for nums in history_df['numbers']]
        self.draw_map = {row['draw_id']: row['numbers'] for _, row in history_df.iterrows()}
        
        self.number_to_draws_index = defaultdict(set)
        for idx, draw_set in enumerate(self.past_draws_sets):
            for num in draw_set:
                self.number_to_draws_index[num].add(idx)
        
        all_numbers = list(chain.from_iterable(history_df['numbers']))
        self.frequency = Counter(all_numbers)
        self.total_draws = len(history_df)
        
        all_sums = [sum(nums) for nums in history_df['numbers']]
        self.global_avg_sum = sum(all_sums) / len(all_sums) if all_sums else 0
        
        sorted_nums = sorted(range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1), 
                           key=lambda x: self.frequency[x], reverse=True)
        self.hot_pool = set(sorted_nums[:16])
        self.cold_pool = set(sorted_nums[16:])
        
        self._calculate_gaps()
        self._analyze_combinations()
    
    def _calculate_gaps(self):
        self.gaps = {}
        for num in range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1):
            appearances = []
            for idx, nums in enumerate(self.history_df['numbers']):
                if num in nums:
                    appearances.append(idx)
            
            if appearances:
                gaps_list = np.diff(appearances) if len(appearances) > 1 else []
                self.gaps[num] = {
                    'last_seen': self.total_draws - 1 - appearances[-1],
                    'avg_gap': np.mean(gaps_list) if len(gaps_list) > 0 else 0,
                    'max_gap': int(np.max(gaps_list)) if len(gaps_list) > 0 else 0,
                    'min_gap': int(np.min(gaps_list)) if len(gaps_list) > 0 else 0,
                    'total_appearances': len(appearances),
                    'appearance_rate': len(appearances) / self.total_draws
                }
            else:
                self.gaps[num] = {
                    'last_seen': self.total_draws,
                    'avg_gap': 0,
                    'max_gap': 0,
                    'min_gap': 0,
                    'total_appearances': 0,
                    'appearance_rate': 0
                }
    
    def _analyze_combinations(self):
        self.pair_freq = Counter()
        self.triple_freq = Counter()
        
        for nums in self.history_df['numbers']:
            for pair in combinations(nums, 2):
                self.pair_freq[tuple(sorted(pair))] += 1
            for triple in combinations(nums, 3):
                self.triple_freq[tuple(sorted(triple))] += 1
    
    def get_number_stats(self, num: int) -> Dict:
        return {
            'frequency': self.frequency.get(num, 0),
            'percentage': (self.frequency.get(num, 0) / (self.total_draws * 6)) * 100,
            'gap_info': self.gaps.get(num, {}),
            'category': 'hot' if num in self.hot_pool else 'cold'
        }
    
    def calculate_ticket_score(self, ticket: List[int]) -> Dict:
        """حساب نقاط التذكرة بناءً على التحليل العميق"""
        ticket_set = set(ticket)
        
        # 1. التحليل الزمني - توزيع التكرارات
        freq_scores = [self.frequency.get(n, 0) for n in ticket]
        temporal_score = np.std(freq_scores) / (np.mean(freq_scores) + 1)
        
        # 2. تحليل الارتباط - الأزواج والثلاثيات الشائعة
        correlation_score = 0
        pairs_count = 0
        for pair in combinations(ticket, 2):
            if tuple(sorted(pair)) in self.pair_freq:
                correlation_score += self.pair_freq[tuple(sorted(pair))]
                pairs_count += 1
        
        triples_count = 0
        for triple in combinations(ticket, 3):
            if tuple(sorted(triple)) in self.triple_freq:
                correlation_score += self.triple_freq[tuple(sorted(triple))] * 2
                triples_count += 1
        
        correlation_score = correlation_score / (pairs_count + triples_count + 1)
        
        # 3. تحليل الفجوات
        gaps = [self.gaps.get(n, {}).get('last_seen', 999) for n in ticket]
        gap_score = np.mean(gaps)
        
        # 4. تحليل الانحراف
        deviation_score = abs(sum(ticket) - self.global_avg_sum) / self.global_avg_sum
        
        # حساب النقاط الإجمالية (0-100)
        total_score = (
            (1 - temporal_score) * 25 +
            (correlation_score / 10) * 25 +
            (gap_score / 50) * 25 +
            (1 - deviation_score) * 25
        )
        
        return {
            'total_score': min(100, max(0, total_score)),
            'temporal': temporal_score,
            'correlation': correlation_score,
            'gap': gap_score,
            'deviation': deviation_score,
            'pairs_found': pairs_count,
            'triples_found': triples_count
        }


class SmartGenerator:
    def __init__(self, analyzer: AdvancedAnalyzer):
        self.analyzer = analyzer
    
    def generate_ticket(
        self,
        ticket_size: int = 6,
        strategy: str = 'balanced',
        sum_target: Optional[int] = None,
        sum_tolerance: float = 0.15,
        consecutive_mode: str = 'allow',  # 'allow', 'avoid', 'require'
        min_consecutives: int = 1,
        max_shadow: int = 3,
        match_last_draw: Optional[int] = None,  # عدد الأرقام المطابقة لآخر سحب (None = لا يهم)
        fixed_numbers: Optional[Set[int]] = None
    ) -> Optional[List[int]]:
        """
        توليد تذكرة ذكية مع استرخاء تدريجي للشروط (مستوحى من Gemini)
        
        الاستراتيجية:
        1. محاولة صارمة للشروط (15000 محاولة)
        2. استرخاء تدريجي لقيد الظلال (±1)
        3. لا يفشل أبداً - يولد تذكرة دائماً
        
        المعاملات:
        - max_shadow: عدد الظلال (تطابق خانة الآحاد)
        - match_last_draw: عدد الأرقام المطابقة لآخر سحب (1-4 أو None)
        
        consecutive_mode:
        - 'allow': يسمح بوجود أو عدم وجود متتاليات
        - 'avoid': يتجنب المتتاليات
        - 'require': يتطلب وجود min_consecutives من المتتاليات على الأقل
        """
        
        fixed_numbers = fixed_numbers or set()
        remaining_size = ticket_size - len(fixed_numbers)
        
        if remaining_size < 0:
            return None
        
        # تحديد المجموعة المتاحة
        if strategy == 'hot':
            pool = list(self.analyzer.hot_pool - fixed_numbers)
        elif strategy == 'cold':
            pool = list(self.analyzer.cold_pool - fixed_numbers)
        else:
            pool = list(set(range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1)) - fixed_numbers)
        
        if len(pool) < remaining_size:
            return None
        
        # المرحلة 1: محاولات صارمة
        for attempt in range(LotteryConfig.STRICT_SHADOW_ATTEMPTS):
            candidate = list(fixed_numbers) + random.sample(pool, remaining_size)
            candidate.sort()
            
            # فحص المتتاليات
            has_consec = self._has_consecutive(candidate)
            consec_count = self._count_consecutives(candidate)
            
            if consecutive_mode == 'avoid' and has_consec:
                continue
            elif consecutive_mode == 'require' and consec_count < min_consecutives:
                continue
            
            # فحص المتوسط
            if sum_target is not None:
                current_sum = sum(candidate)
                tolerance_range = sum_target * sum_tolerance
                if not (sum_target - tolerance_range <= current_sum <= sum_target + tolerance_range):
                    continue
            
            # فحص الظلال (تطابق خانة الآحاد)
            shadow_count = self._count_shadows(candidate)
            if shadow_count != max_shadow:  # دقيق تماماً
                continue
            
            # فحص التطابق مع آخر سحب
            if match_last_draw is not None:
                last_draw_matches = self._count_matches_with_last_draw(candidate)
                if last_draw_matches != match_last_draw:
                    continue
            
            return candidate
        
        # المرحلة 2: استرخاء بسيط لقيد الظلال (±1)
        for attempt in range(LotteryConfig.MAX_GENERATION_ATTEMPTS - LotteryConfig.STRICT_SHADOW_ATTEMPTS):
            candidate = list(fixed_numbers) + random.sample(pool, remaining_size)
            candidate.sort()
            
            # فحص المتتاليات
            has_consec = self._has_consecutive(candidate)
            consec_count = self._count_consecutives(candidate)
            
            if consecutive_mode == 'avoid' and has_consec:
                continue
            elif consecutive_mode == 'require' and consec_count < min_consecutives:
                continue
            
            # فحص المتوسط
            if sum_target is not None:
                current_sum = sum(candidate)
                tolerance_range = sum_target * sum_tolerance
                if not (sum_target - tolerance_range <= current_sum <= sum_target + tolerance_range):
                    continue
            
            # فحص الظلال
            shadow_count = self._count_shadows(candidate)
            # استرخاء: ±1 من الحد المطلوب
            if not (max(0, max_shadow - 1) <= shadow_count <= max_shadow + 1):
                continue
            
            # فحص التطابق مع آخر سحب (استرخاء ±1)
            if match_last_draw is not None:
                last_draw_matches = self._count_matches_with_last_draw(candidate)
                if not (max(0, match_last_draw - 1) <= last_draw_matches <= match_last_draw + 1):
                    continue
            
            return candidate
        
        # المرحلة 3: توليد بدون قيد الظلال فقط (لضمان النجاح)
        for attempt in range(10000):
            candidate = list(fixed_numbers) + random.sample(pool, remaining_size)
            candidate.sort()
            
            # فحص المتتاليات (نحافظ على هذا الشرط)
            has_consec = self._has_consecutive(candidate)
            consec_count = self._count_consecutives(candidate)
            
            if consecutive_mode == 'avoid' and has_consec:
                continue
            elif consecutive_mode == 'require' and consec_count < min_consecutives:
                continue
            
            # فحص المتوسط
            if sum_target is not None:
                current_sum = sum(candidate)
                tolerance_range = sum_target * sum_tolerance
                if not (sum_target - tolerance_range <= current_sum <= sum_target + tolerance_range):
                    continue
            
            shadow_count = self._count_shadows(candidate)
            # استرخاء: ±1 من الحد المطلوب
            if not (max(0, max_shadow - 1) <= shadow_count <= max_shadow + 1):
                continue
            
            return candidate
        
        # المرحلة 3: توليد بدون قيد الظلال فقط (لضمان النجاح)
        for attempt in range(10000):
            candidate = list(fixed_numbers) + random.sample(pool, remaining_size)
            candidate.sort()
            
            # فحص المتتاليات (نحافظ على هذا الشرط)
            has_consec = self._has_consecutive(candidate)
            consec_count = self._count_consecutives(candidate)
            
            if consecutive_mode == 'avoid' and has_consec:
                continue
            elif consecutive_mode == 'require' and consec_count < min_consecutives:
                continue
            
            # فحص المتوسط
            if sum_target is not None:
                current_sum = sum(candidate)
                tolerance_range = sum_target * sum_tolerance
                if not (sum_target - tolerance_range <= current_sum <= sum_target + tolerance_range):
                    continue
            
            # بدون قيد الظلال
            return candidate
        
        # المرحلة 4: توليد أساسي (نادر جداً) - نرخي شرط المتتاليات إذا كان require
        for attempt in range(5000):
            candidate = list(fixed_numbers) + random.sample(pool, remaining_size)
            candidate.sort()
            
            # في المرحلة الأخيرة، نقبل أي شيء ما عدا avoid
            if consecutive_mode == 'avoid':
                has_consec = self._has_consecutive(candidate)
                if has_consec:
                    continue
            
            return candidate
        
        # المرحلة النهائية: توليد بدون أي قيود
        candidate = list(fixed_numbers) + random.sample(pool, remaining_size)
        candidate.sort()
        return candidate
    
    
    def _has_consecutive(self, nums: List[int]) -> bool:
        """فحص وجود متتاليات في التذكرة"""
        for i in range(len(nums) - 1):
            if nums[i + 1] - nums[i] == 1:
                return True
        return False
    
    def _count_consecutives(self, nums: List[int]) -> int:
        """عد عدد المتتاليات في التذكرة"""
        count = 0
        for i in range(len(nums) - 1):
            if nums[i + 1] - nums[i] == 1:
                count += 1
        return count
    
    def _count_shadows(self, nums: List[int]) -> int:
        """
        حساب عدد الظلال في التذكرة
        الظل = مجموعة أرقام لها نفس خانة الآحاد
        مثال: 7 و 27 لهما نفس الآحاد (7) = ظل واحد
        """
        # نحسب خانة الآحاد لكل رقم
        ones_digits = [num % 10 for num in nums]
        # نحسب كم خانة آحاد مكررة
        from collections import Counter
        ones_counter = Counter(ones_digits)
        # عدد الظلال = عدد خانات الآحاد التي تكررت (ظهرت أكثر من مرة)
        shadows_count = sum(1 for count in ones_counter.values() if count > 1)
        return shadows_count
    
    def _count_matches_with_last_draw(self, nums: List[int]) -> int:
        """
        حساب عدد التطابقات مع آخر سحب
        """
        if not self.analyzer.past_draws_sets:
            return 0
        
        ticket_set = set(nums)
        last_draw = self.analyzer.past_draws_sets[-1]  # آخر سحب
        matches = len(ticket_set & last_draw)
        return matches


class DrawSimulator:
    def __init__(self, analyzer: AdvancedAnalyzer):
        self.analyzer = analyzer
    
    def simulate_draws(
        self,
        num_simulations: int,
        ticket: List[int],
        progress_callback=None
    ) -> Dict:
        
        ticket_set = set(ticket)
        results = {3: 0, 4: 0, 5: 0, 6: 0}
        
        for i in range(num_simulations):
            simulated_draw = set(random.sample(range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1), 6))
            matches = len(ticket_set & simulated_draw)
            
            if matches >= 3:
                results[matches] += 1
            
            if progress_callback and (i + 1) % 1000 == 0:
                progress_callback(i + 1, num_simulations)
        
        theoretical_probs = ProbabilityCalculator.calculate_match_probabilities(len(ticket))
        
        output = {
            'simulations': num_simulations,
            'ticket': ticket,
            'results': {}
        }
        
        for match_level in [3, 4, 5, 6]:
            count = results[match_level]
            percentage = (count / num_simulations) * 100
            theoretical = theoretical_probs[match_level] * 100
            
            output['results'][match_level] = {
                'count': count,
                'percentage': percentage,
                'theoretical': theoretical
            }
        
        return output


class ProbabilityCalculator:
    @staticmethod
    def nCr(n: int, r: int) -> int:
        if r > n or r < 0:
            return 0
        if r == 0 or r == n:
            return 1
        
        r = min(r, n - r)
        result = 1
        for i in range(r):
            result = result * (n - i) // (i + 1)
        return result
    
    @staticmethod
    def calculate_match_probabilities(ticket_size: int) -> Dict[int, float]:
        total_numbers = 32
        draw_size = 6
        
        total_combinations = ProbabilityCalculator.nCr(total_numbers, draw_size)
        
        probabilities = {}
        
        for match in range(draw_size + 1):
            ways_to_match = ProbabilityCalculator.nCr(ticket_size, match)
            ways_to_not_match = ProbabilityCalculator.nCr(total_numbers - ticket_size, draw_size - match)
            
            favorable = ways_to_match * ways_to_not_match
            probability = favorable / total_combinations if total_combinations > 0 else 0
            
            probabilities[match] = probability
        
        return probabilities
    
    @staticmethod
    def calculate_expected_value(ticket_size: int, jackpot: int = 1000000) -> Dict:
        probabilities = ProbabilityCalculator.calculate_match_probabilities(ticket_size)
        
        prizes = {
            3: LotteryConfig.MATCH_PRIZES[3],
            4: LotteryConfig.MATCH_PRIZES[4],
            5: LotteryConfig.MATCH_PRIZES[5],
            6: jackpot
        }
        
        expected_value = 0
        breakdown = {}
        
        for match_level in [3, 4, 5, 6]:
            prob = probabilities.get(match_level, 0)
            prize = prizes[match_level]
            contribution = prob * prize
            expected_value += contribution
            
            breakdown[f'{match_level} أرقام'] = {
                'probability': prob,
                'prize': prize,
                'contribution': contribution
            }
        
        ticket_cost = LotteryConfig.TICKET_PRICES.get(ticket_size, 0)
        net_expected_value = expected_value - ticket_cost
        roi = (net_expected_value / ticket_cost * 100) if ticket_cost > 0 else 0
        
        return {
            'ticket_cost': ticket_cost,
            'expected_value': expected_value,
            'net_expected_value': net_expected_value,
            'roi': roi,
            'breakdown': breakdown
        }


class ExportManager:
    @staticmethod
    def export_to_excel(tickets: List[List[int]], filename: str = "tickets.xlsx") -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "التذاكر"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['#'] + [f'N{i}' for i in range(1, 11)] + ['المجموع', 'فردي', 'زوجي']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for idx, ticket in enumerate(tickets, 1):
            ticket_padded = ticket + [None] * (10 - len(ticket))
            row_data = [idx] + ticket_padded + [
                sum(ticket),
                sum(1 for n in ticket if n % 2 == 1),
                sum(1 for n in ticket if n % 2 == 0)
            ]
            ws.append(row_data)
            
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

# ==============================================================================
# 5. واجهات المستخدم
# ==============================================================================

def render_dashboard(analyzer: AdvancedAnalyzer):
    st.header("📊 لوحة المعلومات")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("إجمالي السحوبات", f"{analyzer.total_draws:,}")
    
    with col2:
        most_common = analyzer.frequency.most_common(1)[0]
        st.metric("الأكثر تكراراً", f"{most_common[0]} ({most_common[1]})")
    
    with col3:
        st.metric("متوسط المجموع", f"{analyzer.global_avg_sum:.1f}")
    
    with col4:
        least_common = analyzer.frequency.most_common()[-1]
        st.metric("الأقل تكراراً", f"{least_common[0]} ({least_common[1]})")
    
    st.divider()
    
    col_chart1, col_chart2 = st.columns(2)
    
    with col_chart1:
        st.subheader("📈 توزيع الأرقام")
        freq_df = pd.DataFrame([
            {'رقم': num, 'تكرار': analyzer.frequency.get(num, 0)}
            for num in range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1)
        ])
        
        fig = px.bar(freq_df, x='رقم', y='تكرار', 
                     color='تكرار',
                     color_continuous_scale='Viridis')
        fig.update_layout(showlegend=False, height=400)
        st.plotly_chart(fig, use_container_width=True)
    
    with col_chart2:
        st.subheader("🔥 خريطة حرارية")
        
        matrix_data = []
        for i in range(4):
            row = []
            for j in range(8):
                num = i * 8 + j + 1
                if num <= 32:
                    row.append(analyzer.frequency.get(num, 0))
                else:
                    row.append(0)
            matrix_data.append(row)
        
        fig = go.Figure(data=go.Heatmap(
            z=matrix_data,
            colorscale='RdYlGn',
            showscale=True
        ))
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)
    
    st.divider()
    
    st.subheader("🎲 آخر 5 سحوبات")
    recent_draws = analyzer.history_df.tail(5)[['draw_id', 'numbers', 'sum']].iloc[::-1]
    
    for _, row in recent_draws.iterrows():
        cols = st.columns([1, 6, 1])
        with cols[0]:
            st.write(f"**#{row['draw_id']}**")
        with cols[1]:
            numbers_html = ''.join([
                f'<span class="lottery-number">{num}</span>'
                for num in row['numbers']
            ])
            st.markdown(numbers_html, unsafe_allow_html=True)
        with cols[2]:
            st.write(f"Σ={row['sum']}")


def render_smart_generator(analyzer: AdvancedAnalyzer, generator: SmartGenerator):
    st.header("🎰 المولد الذكي")
    
    col_settings, col_output = st.columns([1, 1])
    
    with col_settings:
        st.subheader("⚙️ الإعدادات")
        
        ticket_size = st.slider(
            "حجم التذكرة",
            LotteryConfig.MIN_TICKET_SIZE,
            LotteryConfig.MAX_TICKET_SIZE,
            LotteryConfig.DEFAULT_TICKET_SIZE
        )
        
        strategy = st.selectbox(
            "الاستراتيجية",
            ['balanced', 'hot', 'cold'],
            format_func=lambda x: {'balanced': '⚖️ متوازنة', 'hot': '🔥 ساخنة', 'cold': '❄️ باردة'}[x]
        )
        
        use_sum_target = st.checkbox("استهداف متوسط محدد")
        sum_target = None
        sum_tolerance = LotteryConfig.DEFAULT_SUM_TOLERANCE
        
        if use_sum_target:
            sum_target = st.number_input(
                "المتوسط المستهدف",
                min_value=ticket_size * LotteryConfig.MIN_NUM,
                max_value=ticket_size * LotteryConfig.MAX_NUM,
                value=int(analyzer.global_avg_sum)
            )
            sum_tolerance = st.slider("التسامح (%)", 5, 30, 15) / 100
        
        consecutive_mode = st.radio(
            "إعداد المتتاليات",
            options=['allow', 'avoid', 'require'],
            format_func=lambda x: {
                'allow': '✅ السماح (غير مهم)',
                'avoid': '🚫 تجنب المتتاليات',
                'require': '✔️ مطلوب متتاليات'
            }[x],
            index=0
        )
        
        min_consecutives = 1
        if consecutive_mode == 'require':
            min_consecutives = st.slider(
                "الحد الأدنى للمتتاليات",
                min_value=1,
                max_value=3,
                value=1,
                help="عدد المتتاليات المطلوبة على الأقل في التذكرة"
            )
        
        max_shadow = st.slider(
            "الحد الأقصى للظلال (تطابق خانة الآحاد)", 
            0, 5, 1,
            help="عدد خانات الآحاد المتكررة. مثال: 7 و27 = ظل واحد"
        )
        
        # خيار التطابق مع آخر سحب
        use_last_draw_match = st.checkbox("التطابق مع آخر سحب", value=False)
        match_last_draw = None
        if use_last_draw_match:
            match_last_draw = st.slider(
                "عدد الأرقام المطابقة لآخر سحب",
                min_value=1,
                max_value=4,
                value=2,
                help="عدد الأرقام من التذكرة التي يجب أن تطابق آخر سحب"
            )
            
            # عرض آخر سحب للمستخدم
            if analyzer.history_df is not None and len(analyzer.history_df) > 0:
                last_draw_numbers = analyzer.history_df.iloc[-1]['numbers']
                st.info(f"📋 آخر سحب: {', '.join(map(str, last_draw_numbers))}")
        
        st.divider()
        st.subheader("📌 أرقام ثابتة (اختياري)")
        fixed_input = st.text_input(
            "أرقام مفصولة بفواصل",
            placeholder="مثال: 5, 12, 23"
        )
        
        fixed_numbers = set()
        if fixed_input.strip():
            try:
                fixed_numbers = set(int(x.strip()) for x in fixed_input.split(','))
                if not all(LotteryConfig.MIN_NUM <= n <= LotteryConfig.MAX_NUM for n in fixed_numbers):
                    st.error("الأرقام يجب أن تكون بين 1 و 32")
                    fixed_numbers = set()
            except:
                st.error("صيغة غير صحيحة")
        
        num_tickets = st.slider("عدد التذاكر", 1, LotteryConfig.MAX_BATCH_SIZE, 1)
        
        # نصائح للمستخدم
        with st.expander("💡 نصائح"):
            st.markdown("""
            ### 🎯 الظلال (تطابق خانة الآحاد):
            - **الظل** = مجموعة أرقام لها نفس خانة الآحاد
            - مثال: `7, 17, 27` → ظل واحد (كلهم آحادهم 7)
            - مثال: `8, 18` و `5, 15, 25` → ظلان
            - **الحد الأقصى = 1**: تذكرة واحدة فقط بها تطابق آحاد
            
            ### 🎲 التطابق مع آخر سحب:
            - عدد الأرقام من التذكرة التي يجب أن تطابق آخر سحب
            - مثال: إذا اخترت 2، ستحتوي التذكرة على رقمين من آخر سحب
            
            ### ✔️ إعداد المتتاليات:
            - *السماح*: لا يهم وجود أو عدم وجود متتاليات
            - *تجنب*: لن تحتوي التذاكر على أرقام متتالية (مثل 5,6)
            - *مطلوب*: كل تذكرة ستحتوي على متتالية واحدة على الأقل (مثل 5,6 أو 18,19)
            
            ### ⚙️ التوليد:
            - **الحد الأقصى للظلال = الحد الدقيق**: المولد سيحاول إيجاد تذاكر بهذا الحد بالضبط
            - **استرخاء تلقائي**: إذا لم يجد، سيسمح بـ ±1 من الحد المطلوب
            - **نجاح مضمون**: المولد لن يفشل، سيولد تذاكر دائماً
            """)
        
        if st.button("🎲 توليد", type="primary", use_container_width=True):
            with st.spinner("🔍 جاري البحث عن التذاكر..."):
                tickets = []
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for i in range(num_tickets):
                    status_text.text(f"⏳ توليد تذكرة {i+1} من {num_tickets}...")
                    
                    ticket = generator.generate_ticket(
                        ticket_size=ticket_size,
                        strategy=strategy,
                        sum_target=sum_target,
                        sum_tolerance=sum_tolerance,
                        consecutive_mode=consecutive_mode,
                        min_consecutives=min_consecutives,
                        max_shadow=max_shadow,
                        match_last_draw=match_last_draw,
                        fixed_numbers=fixed_numbers
                    )
                    
                    if ticket:
                        tickets.append(ticket)
                        status_text.text(f"✅ تذكرة {i+1} جاهزة!")
                    else:
                        status_text.text(f"⚠️ لم يتم العثور على تذكرة {i+1}")
                    
                    progress_bar.progress((i + 1) / num_tickets)
                
                progress_bar.empty()
                status_text.empty()
                
                if tickets:
                    st.session_state['generated_tickets_session'] = tickets
                    
                    # رسالة نجاح مع معلومات إضافية
                    success_msg = f"✅ تم توليد {len(tickets)} تذكرة بنجاح"
                    if len(tickets) < num_tickets:
                        success_msg += f" (من أصل {num_tickets} مطلوبة)"
                    st.success(success_msg)
                    
                    # عرض إحصائيات التذاكر المولدة
                    if tickets:
                        avg_sum = sum(sum(t) for t in tickets) / len(tickets)
                        avg_shadows = sum(generator._count_shadows(t) for t in tickets) / len(tickets)
                        st.info(f"""
                        📊 **إحصائيات التذاكر:**
                        - متوسط المجموع: {avg_sum:.1f}
                        - متوسط الظلال: {avg_shadows:.1f}
                        """)
                else:
                    st.error("⚠️ فشل التوليد")
                    st.warning("""
                    **الشروط صعبة جداً. جرب:**
                    - زيادة الحد الأقصى للظلال بمقدار 1
                    - تغيير إعداد المتتاليات إلى "السماح"
                    - زيادة نسبة التسامح
                    """)
    
    with col_output:
        if st.session_state.get('generated_tickets_session'):
            st.subheader("🎟️ التذاكر المولدة")
            
            for idx, ticket in enumerate(st.session_state['generated_tickets_session'], 1):
                with st.container():
                    st.markdown(f"**تذكرة #{idx}**")
                    numbers_html = ''.join([
                        f'<span class="lottery-number">{num}</span>'
                        for num in ticket
                    ])
                    st.markdown(numbers_html, unsafe_allow_html=True)
                    
                    # حساب عدد المتتاليات
                    consec_count = generator._count_consecutives(ticket)
                    consec_text = f"متتاليات: {consec_count}" if consec_count > 0 else "لا متتاليات"
                    
                    # حساب عدد الظلال
                    shadow_count = generator._count_shadows(ticket)
                    shadow_text = f"ظلال: {shadow_count}"
                    
                    # حساب التطابق مع آخر سحب
                    last_draw_match = generator._count_matches_with_last_draw(ticket)
                    match_text = f"تطابق آخر سحب: {last_draw_match}"
                    
                    st.caption(f"المجموع: {sum(ticket)} | {consec_text} | {shadow_text} | {match_text} | السعر: {LotteryConfig.TICKET_PRICES[len(ticket)]} د")
                    st.divider()
            
            excel_data = ExportManager.export_to_excel(st.session_state['generated_tickets_session'])
            st.download_button(
                "📥 تحميل Excel",
                data=excel_data,
                file_name=f"tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )


def render_advanced_filter(analyzer: AdvancedAnalyzer, generator: SmartGenerator):
    """نافذة ترشيح متقدمة للتذاكر بناءً على التحليل العميق"""
    st.header("🔬 ترشيح التذاكر المتقدم")
    
    st.info("""
    ### 📋 كيف تعمل؟
    هذه الأداة تولد تذاكر متعددة ثم تقوم بفرزها وترشيحها بناءً على:
    - **التحليل الزمني**: توزيع تكرار الأرقام
    - **تحليل الارتباط**: الأزواج والثلاثيات الشائعة
    - **تحليل الفجوات**: آخر ظهور للأرقام
    - **تحليل الانحراف**: مدى قرب المجموع من المتوسط
    """)
    
    st.divider()
    
    col_filter1, col_filter2 = st.columns([1, 1])
    
    with col_filter1:
        st.subheader("⚙️ إعدادات التوليد")
        
        filter_ticket_size = st.selectbox(
            "حجم التذاكر",
            [6, 7, 8, 9, 10],
            index=0
        )
        
        filter_strategy = st.selectbox(
            "الاستراتيجية",
            ['balanced', 'hot', 'cold'],
            format_func=lambda x: {'balanced': '⚖️ متوازنة', 'hot': '🔥 ساخنة', 'cold': '❄️ باردة'}[x],
            key='filter_strategy'
        )
        
        num_candidates = st.slider(
            "عدد التذاكر المرشحة",
            min_value=50,
            max_value=500,
            value=100,
            step=50,
            help="عدد التذاكر التي سيتم توليدها قبل الترشيح"
        )
        
        top_n = st.slider(
            "أفضل N تذكرة",
            min_value=5,
            max_value=50,
            value=10,
            help="عدد أفضل التذاكر التي سيتم عرضها"
        )
        
    with col_filter2:
        st.subheader("🎯 معايير الترشيح")
        
        weight_temporal = st.slider(
            "وزن التحليل الزمني",
            min_value=0.0,
            max_value=1.0,
            value=0.25,
            step=0.05,
            help="أهمية التجانس في توزيع تكرارات الأرقام"
        )
        
        weight_correlation = st.slider(
            "وزن الارتباط",
            min_value=0.0,
            max_value=1.0,
            value=0.25,
            step=0.05,
            help="أهمية وجود أزواج/ثلاثيات شائعة"
        )
        
        weight_gap = st.slider(
            "وزن الفجوات",
            min_value=0.0,
            max_value=1.0,
            value=0.25,
            step=0.05,
            help="أهمية الأرقام التي لم تظهر منذ فترة"
        )
        
        weight_deviation = st.slider(
            "وزن الانحراف",
            min_value=0.0,
            max_value=1.0,
            value=0.25,
            step=0.05,
            help="أهمية قرب المجموع من المتوسط"
        )
    
    st.divider()
    
    if st.button("🚀 توليد وترشيح", type="primary", use_container_width=True):
        with st.spinner(f"جاري توليد {num_candidates} تذكرة وتحليلها..."):
            candidate_tickets = []
            progress_bar = st.progress(0)
            progress_text = st.empty()
            
            for i in range(num_candidates):
                ticket = generator.generate_ticket(
                    ticket_size=filter_ticket_size,
                    strategy=filter_strategy,
                    consecutive_mode='allow',
                    max_shadow=3
                )
                
                if ticket:
                    candidate_tickets.append(ticket)
                
                if (i + 1) % 10 == 0:
                    progress_bar.progress((i + 1) / num_candidates)
                    progress_text.text(f"تم توليد {i + 1} / {num_candidates} تذكرة")
            
            progress_bar.empty()
            progress_text.empty()
            
            if not candidate_tickets:
                st.error("فشل توليد التذاكر. جرب إعدادات مختلفة.")
                return
            
            st.info(f"تم توليد {len(candidate_tickets)} تذكرة بنجاح. جاري التحليل...")
            
            scored_tickets = []
            for ticket in candidate_tickets:
                score_data = analyzer.calculate_ticket_score(ticket)
                
                weighted_score = (
                    (1 - score_data['temporal']) * weight_temporal * 100 +
                    (score_data['correlation'] / 10) * weight_correlation * 100 +
                    (score_data['gap'] / 50) * weight_gap * 100 +
                    (1 - score_data['deviation']) * weight_deviation * 100
                )
                
                scored_tickets.append({
                    'ticket': ticket,
                    'score': weighted_score,
                    'details': score_data
                })
            
            scored_tickets.sort(key=lambda x: x['score'], reverse=True)
            top_tickets = scored_tickets[:top_n]
            
            st.session_state['filtered_tickets'] = top_tickets
            
            st.success(f"✅ تم ترشيح أفضل {len(top_tickets)} تذكرة!")
    
    if st.session_state.get('filtered_tickets'):
        st.divider()
        st.subheader(f"🏆 أفضل {len(st.session_state['filtered_tickets'])} تذكرة")
        
        for idx, item in enumerate(st.session_state['filtered_tickets'], 1):
            ticket = item['ticket']
            score = item['score']
            details = item['details']
            
            if score >= 80:
                badge_color = "#22c55e"
            elif score >= 60:
                badge_color = "#f59e0b"
            else:
                badge_color = "#3b82f6"
            
            with st.container():
                st.markdown(f"""
                <div class="filter-card">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <h4>تذكرة #{idx}</h4>
                        <span class="score-badge" style="background-color: {badge_color}; color: white;">
                            النقاط: {score:.1f}/100
                        </span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                
                numbers_html = ''.join([
                    f'<span class="lottery-number">{num}</span>'
                    for num in ticket
                ])
                st.markdown(numbers_html, unsafe_allow_html=True)
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("المجموع", sum(ticket))
                with col2:
                    st.metric("أزواج شائعة", details['pairs_found'])
                with col3:
                    st.metric("ثلاثيات شائعة", details['triples_found'])
                with col4:
                    st.metric("السعر", f"{LotteryConfig.TICKET_PRICES[len(ticket)]} د")
                
                with st.expander("📊 التفاصيل الكاملة"):
                    st.json({
                        'التجانس الزمني': f"{details['temporal']:.4f}",
                        'الارتباط': f"{details['correlation']:.2f}",
                        'متوسط الفجوة': f"{details['gap']:.1f}",
                        'الانحراف': f"{details['deviation']:.4f}"
                    })
                
                st.divider()
        
        tickets_to_export = [item['ticket'] for item in st.session_state['filtered_tickets']]
        excel_data = ExportManager.export_to_excel(tickets_to_export)
        st.download_button(
            "📥 تحميل أفضل التذاكر (Excel)",
            data=excel_data,
            file_name=f"filtered_tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )


def render_checker(analyzer: AdvancedAnalyzer):
    st.header("🔍 فاحص التذاكر")
    
    ticket_input = st.text_input(
        "أدخل أرقامك (مفصولة بفواصل)",
        placeholder="مثال: 5, 12, 18, 23, 27, 31"
    )
    
    if ticket_input:
        try:
            ticket = sorted([int(x.strip()) for x in ticket_input.split(',')])
            
            if not all(LotteryConfig.MIN_NUM <= n <= LotteryConfig.MAX_NUM for n in ticket):
                st.error("جميع الأرقام يجب أن تكون بين 1 و 32")
                return
            
            if len(ticket) != len(set(ticket)):
                st.error("الأرقام يجب أن تكون فريدة")
                return
            
            numbers_html = ''.join([
                f'<span class="lottery-number">{num}</span>'
                for num in ticket
            ])
            st.markdown(numbers_html, unsafe_allow_html=True)
            
            st.divider()
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("📊 الإحصائيات")
                st.metric("المجموع", sum(ticket))
                st.metric("الأرقام الفردية", sum(1 for n in ticket if n % 2 == 1))
                st.metric("الأرقام الزوجية", sum(1 for n in ticket if n % 2 == 0))
                
                has_consecutive = False
                for i in range(len(ticket) - 1):
                    if ticket[i + 1] - ticket[i] == 1:
                        has_consecutive = True
                        break
                
                st.metric("متتاليات", "نعم" if has_consecutive else "لا")
            
            with col2:
                st.subheader("🔥 التكرار")
                for num in ticket:
                    freq = analyzer.frequency.get(num, 0)
                    percentage = (freq / analyzer.total_draws) * 100
                    category = "🔥" if num in analyzer.hot_pool else "❄️"
                    st.write(f"{category} **{num}**: {freq} مرة ({percentage:.1f}%)")
            
            st.divider()
            
            st.subheader("🎯 التطابقات التاريخية")
            ticket_set = set(ticket)
            matches = []
            
            for idx, row in analyzer.history_df.iterrows():
                past_set = set(row['numbers'])
                match_count = len(ticket_set & past_set)
                
                if match_count >= 3:
                    matches.append({
                        'draw_id': row['draw_id'],
                        'numbers': row['numbers'],
                        'matches': match_count
                    })
            
            if matches:
                matches.sort(key=lambda x: x['matches'], reverse=True)
                
                for match in matches[:10]:
                    st.write(f"**السحب #{match['draw_id']}** - {match['matches']} تطابقات")
                    matched_nums = ticket_set & set(match['numbers'])
                    st.write(f"الأرقام المتطابقة: {sorted(matched_nums)}")
                    st.divider()
            else:
                st.info("لا توجد تطابقات (3+) في السحوبات السابقة")
        
        except:
            st.error("صيغة غير صحيحة. استخدم أرقاماً مفصولة بفواصل")


def render_deep_analytics(analyzer: AdvancedAnalyzer):
    st.header("📈 التحليلات المتقدمة")
    
    analysis_tabs = st.tabs([
        "⏰ زمني",
        "🔗 ارتباط",
        "📊 فجوات",
        "📉 انحراف"
    ])
    
    with analysis_tabs[0]:
        st.subheader("التحليل الزمني")
        
        if 'date' in analyzer.history_df.columns:
            df_time = analyzer.history_df.copy()
            
            # خيارات التحليل الزمني
            time_analysis_type = st.radio(
                "نوع التحليل",
                ["الشهري", "الأسبوعي", "اليومي"],
                horizontal=True
            )
            
            if time_analysis_type == "الشهري":
                df_time['period'] = pd.to_datetime(df_time['date']).dt.to_period('M').astype(str)
            elif time_analysis_type == "الأسبوعي":
                df_time['period'] = pd.to_datetime(df_time['date']).dt.isocalendar().week.astype(str)
                df_time['period'] = "أسبوع " + df_time['period']
            else:
                df_time['period'] = pd.to_datetime(df_time['date']).dt.date.astype(str)
            
            # حساب الإحصائيات مع تجنب MultiIndex
            monthly_stats = df_time.groupby('period').agg({
                'sum': 'mean',
                'odd_count': 'mean',
                'even_count': 'mean'
            }).reset_index()
            
            # رسم المخطط
            fig = px.line(
                monthly_stats,
                x='period',
                y='sum',
                title=f'متوسط المجموع حسب {time_analysis_type}',
                labels={'sum': 'متوسط المجموع', 'period': f'{time_analysis_type}'},
                markers=True
            )
            
            fig.update_layout(
                xaxis_title=f"{time_analysis_type}",
                yaxis_title="متوسط المجموع",
                hovermode='x unified'
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            # إضافة مخططات إضافية
            col1, col2 = st.columns(2)
            
            with col1:
                fig2 = px.bar(
                    monthly_stats,
                    x='period',
                    y='odd_count',
                    title='متوسط الأرقام الفردية',
                    labels={'odd_count': 'متوسط الأرقام الفردية'}
                )
                st.plotly_chart(fig2, use_container_width=True)
            
            with col2:
                fig3 = px.bar(
                    monthly_stats,
                    x='period',
                    y='even_count',
                    title='متوسط الأرقام الزوجية',
                    labels={'even_count': 'متوسط الأرقام الزوجية'},
                    color_discrete_sequence=['#f59e0b']
                )
                st.plotly_chart(fig3, use_container_width=True)
    
    with analysis_tabs[1]:
        st.subheader("تحليل الارتباط")
        
        st.write("**أكثر الأزواج شيوعاً:**")
        top_pairs = analyzer.pair_freq.most_common(10)
        
        # إنشاء DataFrame للعرض
        pair_data = []
        for pair, count in top_pairs:
            percentage = (count / analyzer.total_draws) * 100
            pair_data.append({
                'الرقم 1': pair[0],
                'الرقم 2': pair[1],
                'التكرار': count,
                'النسبة %': f"{percentage:.1f}"
            })
        
        df_pairs = pd.DataFrame(pair_data)
        st.dataframe(df_pairs, hide_index=True, use_container_width=True)
        
        # مخطط الأزواج
        if top_pairs:
            pairs_df = pd.DataFrame([
                {
                    'زوج': f"{p[0]}-{p[1]}",
                    'تكرار': c,
                    'نسبة %': (c / analyzer.total_draws) * 100
                }
                for p, c in top_pairs
            ])
            
            fig = px.bar(
                pairs_df,
                x='زوج',
                y='تكرار',
                title='أكثر الأزواج تكراراً',
                color='تكرار',
                color_continuous_scale='Viridis'
            )
            st.plotly_chart(fig, use_container_width=True)
        
        st.divider()
        
        st.write("**أكثر الثلاثيات شيوعاً:**")
        top_triples = analyzer.triple_freq.most_common(5)
        
        triple_data = []
        for triple, count in top_triples:
            percentage = (count / analyzer.total_draws) * 100
            triple_data.append({
                'الرقم 1': triple[0],
                'الرقم 2': triple[1],
                'الرقم 3': triple[2],
                'التكرار': count,
                'النسبة %': f"{percentage:.1f}"
            })
        
        df_triples = pd.DataFrame(triple_data)
        st.dataframe(df_triples, hide_index=True, use_container_width=True)
    
    with analysis_tabs[2]:
        st.subheader("تحليل الفجوات")
        
        # خيارات الفرز
        sort_by = st.selectbox(
            "فرز حسب",
            ["آخر ظهور (تنازلي)", "متوسط الفجوة", "التكرار"],
            key='gap_sort'
        )
        
        gap_data = []
        for num in range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1):
            gap_info = analyzer.gaps.get(num, {})
            gap_data.append({
                'رقم': num,
                'آخر ظهور': gap_info.get('last_seen', 0),
                'متوسط الفجوة': gap_info.get('avg_gap', 0),
                'أقصى فجوة': gap_info.get('max_gap', 0),
                'أدنى فجوة': gap_info.get('min_gap', 0),
                'مجموع الظهور': gap_info.get('total_appearances', 0),
                'نسبة الظهور %': f"{gap_info.get('appearance_rate', 0) * 100:.1f}"
            })
        
        df_gaps = pd.DataFrame(gap_data)
        
        # تطبيق الفرز
        if sort_by == "آخر ظهور (تنازلي)":
            df_gaps = df_gaps.sort_values('آخر ظهور', ascending=False)
        elif sort_by == "متوسط الفجوة":
            df_gaps = df_gaps.sort_values('متوسط الفجوة', ascending=False)
        else:
            df_gaps = df_gaps.sort_values('مجموع الظهور', ascending=False)
        
        st.dataframe(df_gaps, hide_index=True, use_container_width=True)
        
        # مخطط الفجوات
        fig = px.bar(
            df_gaps.head(15),
            x='رقم',
            y='آخر ظهور',
            title='الأرقام التي لم تظهر منذ أطول فترة (أعلى 15)',
            color='آخر ظهور',
            color_continuous_scale='Reds'
        )
        fig.update_layout(
            xaxis_title="الرقم",
            yaxis_title="عدد السحوبات منذ آخر ظهور"
        )
        st.plotly_chart(fig, use_container_width=True)
        
        # مخطط التوزيع
        fig2 = px.histogram(
            df_gaps,
            x='آخر ظهور',
            nbins=20,
            title='توزيع فترات آخر ظهور',
            labels={'آخر ظهور': 'سحوبات منذ آخر ظهور'}
        )
        st.plotly_chart(fig2, use_container_width=True)
    
    with analysis_tabs[3]:
        st.subheader("تحليل الانحراف")
        
        sums = analyzer.history_df['sum'].values
        mean_sum = np.mean(sums)
        std_sum = np.std(sums)
        median_sum = np.median(sums)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("المتوسط", f"{mean_sum:.1f}")
        with col2:
            st.metric("الوسيط", f"{median_sum:.1f}")
        with col3:
            st.metric("الانحراف المعياري", f"{std_sum:.2f}")
        with col4:
            st.metric("معامل الاختلاف", f"{(std_sum / mean_sum * 100):.1f}%")
        
        col_ranges, col_skew = st.columns(2)
        
        with col_ranges:
            st.metric("الحد الأدنى", int(min(sums)))
            st.metric("الحد الأقصى", int(max(sums)))
            st.metric("المدى", int(max(sums) - min(sums)))
        
        with col_skew:
            skewness = pd.Series(sums).skew()
            kurtosis = pd.Series(sums).kurtosis()
            st.metric("الانحراف", f"{skewness:.2f}")
            st.metric("التفرطح", f"{kurtosis:.2f}")
        
        # مخططات متعددة
        fig_col1, fig_col2 = st.columns(2)
        
        with fig_col1:
            fig = px.histogram(
                x=sums,
                nbins=30,
                title='توزيع المجاميع',
                labels={'x': 'المجموع', 'y': 'التكرار'},
                color_discrete_sequence=['#3b82f6']
            )
            fig.add_vline(x=mean_sum, line_dash="dash", line_color="red", 
                         annotation_text="المتوسط", annotation_position="top right")
            fig.add_vline(x=median_sum, line_dash="dot", line_color="green",
                         annotation_text="الوسيط", annotation_position="top left")
            st.plotly_chart(fig, use_container_width=True)
        
        with fig_col2:
            fig2 = px.box(
                x=sums,
                title='مخطط الصندوق للمجاميع',
                labels={'x': 'المجموع'}
            )
            st.plotly_chart(fig2, use_container_width=True)
        
        # مخطط Q-Q (تقريبي)
        st.subheader("اختبار التوزيع الطبيعي")
        
        # حساب النقاط النظرية للتوزيع الطبيعي
        sorted_sums = np.sort(sums)
        n = len(sorted_sums)
        theoretical_quantiles = np.percentile(np.random.normal(mean_sum, std_sum, 10000), 
                                             np.linspace(0, 100, n))
        
        qq_data = pd.DataFrame({
            'النظرية': theoretical_quantiles,
            'الفعلي': sorted_sums
        })
        
        fig3 = px.scatter(
            qq_data,
            x='النظرية',
            y='الفعلي',
            title='مخطط Q-Q (التوزيع الطبيعي)',
            labels={'النظرية': 'النقاط النظرية', 'الفعلي': 'النقاط الفعلية'}
        )
        
        # إضافة خط y=x للمقارنة
        min_val = min(theoretical_quantiles.min(), sorted_sums.min())
        max_val = max(theoretical_quantiles.max(), sorted_sums.max())
        fig3.add_trace(go.Scatter(
            x=[min_val, max_val],
            y=[min_val, max_val],
            mode='lines',
            line=dict(color='red', dash='dash'),
            name='y=x (التوزيع المثالي)'
        ))
        
        # إضافة خط الانحدار البسيط يدوياً
        from numpy.polynomial import Polynomial
        p = Polynomial.fit(theoretical_quantiles, sorted_sums, 1)
        trend_y = p(theoretical_quantiles)
        fig3.add_trace(go.Scatter(
            x=theoretical_quantiles,
            y=trend_y,
            mode='lines',
            line=dict(color='blue', dash='dot'),
            name='خط الاتجاه'
        ))
        
        st.plotly_chart(fig3, use_container_width=True)


def render_probability_simulator(analyzer: AdvancedAnalyzer):
    st.header("🧮 حاسبة الاحتمالات والمحاكاة")
    
    prob_tabs = st.tabs([
        "📊 الاحتمالات",
        "🎲 المحاكاة",
        "💰 القيمة المتوقعة"
    ])
    
    with prob_tabs[0]:
        st.subheader("احتمالات التطابق")
        
        calc_ticket_size = st.slider(
            "حجم التذكرة",
            LotteryConfig.MIN_TICKET_SIZE,
            LotteryConfig.MAX_TICKET_SIZE,
            LotteryConfig.DEFAULT_TICKET_SIZE,
            key='prob_calc_size'
        )
        
        probabilities = ProbabilityCalculator.calculate_match_probabilities(calc_ticket_size)
        
        prob_data = []
        for match_level in range(7):
            prob = probabilities.get(match_level, 0)
            odds = f"1 في {int(1/prob):,}" if prob > 0 else "مستحيل"
            prob_data.append({
                'التطابق': f'{match_level} أرقام',
                'الاحتمال': f'{prob * 100:.6f}%',
                'الاحتمالية': odds
            })
        
        df_prob = pd.DataFrame(prob_data)
        st.dataframe(df_prob, hide_index=True, use_container_width=True)
    
    with prob_tabs[1]:
        st.subheader("محاكاة السحوبات")
        
        col_sim1, col_sim2 = st.columns([1, 1])
        
        with col_sim1:
            num_sims = st.number_input(
                "عدد المحاكاة",
                min_value=1000,
                max_value=1000000,
                value=10000,
                step=1000
            )
            
            sim_ticket = st.multiselect(
                "اختر 6 أرقام",
                options=list(range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1)),
                max_selections=6,
                default=[5, 12, 18, 23, 27, 31],
                key='sim_ticket_select'
            )
            
            if st.button("🚀 محاكاة", type="primary") and len(sim_ticket) == 6:
                simulator = DrawSimulator(analyzer)
                
                progress_bar = st.progress(0)
                progress_text = st.empty()
                
                def sim_progress(current, total):
                    progress_bar.progress(current / total)
                    progress_text.text(f"{current:,} / {total:,}")
                
                with st.spinner("جاري المحاكاة..."):
                    results = simulator.simulate_draws(num_sims, sorted(sim_ticket), sim_progress)
                
                progress_bar.empty()
                progress_text.empty()
                
                st.success(f"✅ تمت {num_sims:,} محاكاة!")
                st.session_state['sim_results'] = results
        
        with col_sim2:
            if 'sim_results' in st.session_state:
                results = st.session_state['sim_results']
                
                st.markdown("### 📊 النتائج:")
                
                results_data = []
                for match_level in [3, 4, 5, 6]:
                    data = results['results'][match_level]
                    results_data.append({
                        'التطابق': f'{match_level} أرقام',
                        'العدد': data['count'],
                        'النسبة': f"{data['percentage']:.4f}%",
                        'النظرية': f"{data['theoretical']:.4f}%"
                    })
                
                df_results = pd.DataFrame(results_data)
                st.dataframe(df_results, hide_index=True, use_container_width=True)
    
    with prob_tabs[2]:
        st.subheader("💰 القيمة المتوقعة")
        
        col_ev1, col_ev2 = st.columns([1, 1])
        
        with col_ev1:
            ticket_size_ev = st.slider("حجم التذكرة:", 6, 10, 6, key='ev_size')
            jackpot_amount = st.number_input(
                "الجائزة الكبرى:",
                min_value=100000,
                max_value=10000000,
                value=1000000,
                step=100000
            )
            
            ev_data = ProbabilityCalculator.calculate_expected_value(ticket_size_ev, jackpot_amount)
            
            st.metric("التكلفة", f"{ev_data['ticket_cost']} د")
            st.metric("القيمة المتوقعة", f"{ev_data['expected_value']:.2f} د")
            st.metric("الصافي", f"{ev_data['net_expected_value']:.2f} د")
            
            roi_color = "green" if ev_data['roi'] > 0 else "red"
            st.markdown(f"**العائد على الاستثمار:** :{roi_color}[{ev_data['roi']:.2f}%]")
        
        with col_ev2:
            st.markdown("### 🎁 التفصيل:")
            
            breakdown_data = []
            for level, data in ev_data['breakdown'].items():
                prize_display = f"{data['prize']:,}" if isinstance(data['prize'], (int, float)) else data['prize']
                breakdown_data.append({
                    'المستوى': level,
                    'الجائزة': prize_display,
                    'المساهمة': f"{data['contribution']:.4f} د"
                })
            
            df_breakdown = pd.DataFrame(breakdown_data)
            st.dataframe(df_breakdown, hide_index=True, use_container_width=True)


def render_user_guide():
    st.header("📖 دليل الاستخدام")
    
    with st.expander("📊 **Dashboard**", expanded=True):
        st.markdown("""
        ### الوظيفة:
        نظرة سريعة على البيانات
        
        ### المحتوى:
        - إحصائيات سريعة
        - أكثر/أقل الأرقام ظهوراً
        - خريطة حرارية
        - آخر 5 سحوبات
        """)
    
    with st.expander("🎰 **المولد الذكي**"):
        st.markdown("""
        ### الإعدادات:
        - **الاستراتيجية:** ساخنة/باردة/متوازنة
        - **المتوسط:** الالتزام بمتوسط محدد
        - **حجم التذكرة:** 6-10
        - **المتتاليات والظلال**
        - **تثبيت أرقام**
        
        ### الخطوات:
        1. اختر المعايير
        2. فحص الجدوى
        3. توليد
        4. تصدير
        """)
    
    with st.expander("🔬 **الترشيح المتقدم** (جديد!)"):
        st.markdown("""
        ### الوظيفة:
        توليد وترشيح التذاكر بناءً على تحليل عميق
        
        ### المعايير:
        - **التحليل الزمني**: توزيع تكرار الأرقام
        - **الارتباط**: الأزواج والثلاثيات الشائعة
        - **الفجوات**: آخر ظهور للأرقام
        - **الانحراف**: قرب المجموع من المتوسط
        
        ### الخطوات:
        1. اختر حجم التذكرة (6-10)
        2. حدد عدد التذاكر المرشحة
        3. اضبط الأوزان حسب تفضيلاتك
        4. اضغط توليد وترشيح
        5. احصل على أفضل التذاكر مرتبة
        """)
    
    with st.expander("🔍 **الفاحص**"):
        st.markdown("""
        ### الفحوصات:
        - التطابقات التاريخية
        - تكرار الأرقام
        - المتتاليات
        """)
    
    with st.expander("📈 **التحليلات**"):
        st.markdown("""
        - **زمني:** توزيع حسب الشهر/السنة
        - **ارتباط:** أزواج وثلاثيات شائعة
        - **فجوات:** آخر ظهور
        - **انحراف:** قياس التشتت
        """)
    
    with st.expander("🧮 **الاحتمالات**"):
        st.markdown("""
        - حساب احتمالية التطابق
        - محاكاة السحوبات
        - القيمة المتوقعة
        """)
    
    st.divider()
    
    st.warning("""
    ⚠️ **إخلاء مسؤولية:**
    
    هذا تطبيق تحليلي تعليمي. اليانصيب لعبة حظ عشوائية.
    لا توجد خوارزمية تضمن الفوز. العب بمسؤولية!
    """)


# ==============================================================================
# 6. الدالة الرئيسية
# ==============================================================================

def main():
    st.set_page_config(
        page_title="🎰 اليانصيب الأردني",
        page_icon="🎲",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    initialize_session_state()
    apply_theme()
    
    # Sidebar
    with st.sidebar:
        st.title("⚙️ الإعدادات")
        
        theme_label = "🌙 داكن" if st.session_state.theme == 'light' else "☀️ فاتح"
        if st.button(theme_label, use_container_width=True):
            st.session_state.theme = 'dark' if st.session_state.theme == 'light' else 'light'
            st.rerun()
        
        st.divider()
        
        st.subheader("🎨 الألوان")
        st.session_state.hot_color = st.color_picker("ساخنة", st.session_state.hot_color)
        st.session_state.cold_color = st.color_picker("باردة", st.session_state.cold_color)
        st.session_state.balanced_color = st.color_picker("متوازنة", st.session_state.balanced_color)
        
        st.divider()
        
        st.subheader("📂 البيانات")
        
        # التحميل التلقائي من GitHub عند أول دخول
        if not st.session_state.get('auto_loaded') and st.session_state.history_df is None:
            with st.spinner("🔄 جاري تحميل البيانات من GitHub..."):
                df, msg = load_from_github()
                if df is not None:
                    st.session_state.history_df = df
                    st.session_state.analyzer = AdvancedAnalyzer(df)
                    st.session_state.generator = SmartGenerator(st.session_state.analyzer)
                    st.session_state.auto_loaded = True
                    st.success(msg)
                else:
                    st.warning(msg)
                    st.info("يرجى رفع ملف يدوياً")
        
        # زر إعادة تحميل من GitHub
        if st.session_state.history_df is not None and st.session_state.auto_loaded:
            if st.button("🔄 إعادة تحميل من GitHub", use_container_width=True):
                # مسح الكاش لفرض إعادة التحميل
                load_from_github.clear()
                with st.spinner("جاري إعادة التحميل..."):
                    df, msg = load_from_github()
                    if df is not None:
                        st.session_state.history_df = df
                        st.session_state.analyzer = AdvancedAnalyzer(df)
                        st.session_state.generator = SmartGenerator(st.session_state.analyzer)
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
        
        st.divider()
        # خيار الرفع اليدوي
        uploaded_file = st.file_uploader("رفع ملف (اختياري)", type=['xlsx', 'xls', 'csv'])
        
        if uploaded_file:
            with st.spinner("جاري التحميل..."):
                df, msg = load_and_process_data(uploaded_file)
                if df is not None:
                    st.session_state.history_df = df
                    st.session_state.analyzer = AdvancedAnalyzer(df)
                    st.session_state.generator = SmartGenerator(st.session_state.analyzer)
                    st.success(f"✅ {len(df)} سحب")
                else:
                    st.error(msg)
    
    if st.session_state.history_df is None:
        st.warning("⚠️ يرجى التأكد من وجود ملف history.xlsx في نفس المجلد")
        st.info("""
        ### 📋 المتطلبات:
        - ملف Excel بإسم `history.xlsx` في نفس مجلد التطبيق
        - الأعمدة: N1, N2, N3, N4, N5, N6
        - الأرقام: 1-32
        
        أو يمكنك رفع الملف يدوياً من الشريط الجانبي
        """)
        return
    
    analyzer = st.session_state.analyzer
    generator = st.session_state.generator
    
    st.title("🎰 اليانصيب الأردني المتطور")
    st.markdown("**النسخة 3.0** - تحليلات متقدمة + ترشيح ذكي")
    
    tabs = st.tabs([
        "📊 Dashboard",
        "🎰 المولد",
        "🔬 الترشيح المتقدم",
        "🔍 الفاحص",
        "📈 التحليلات",
        "🧮 الاحتمالات",
        "📖 الدليل"
    ])
    
    with tabs[0]:
        render_dashboard(analyzer)
    
    with tabs[1]:
        render_smart_generator(analyzer, generator)
    
    with tabs[2]:
        render_advanced_filter(analyzer, generator)
    
    with tabs[3]:
        render_checker(analyzer)
    
    with tabs[4]:
        render_deep_analytics(analyzer)
    
    with tabs[5]:
        render_probability_simulator(analyzer)
    
    with tabs[6]:
        render_user_guide()
    
    st.markdown("""
    <div class="footer">
        <b>برمجة: محمد العمري</b> | النسخة 3.0.0 | فبراير 2026
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
